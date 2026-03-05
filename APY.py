import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import re

st.set_page_config(
    page_title="KRA Auto-Updater | IFB",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
* { font-family: 'Inter', sans-serif !important; }
.stApp { background: #060B18; }
[data-testid="stHeader"] { background: transparent; }
[data-testid="stSidebar"] { display: none; }
h1,h2,h3,p,label,span,div { color: #E2E8F0; }
.upload-card {
    background: linear-gradient(145deg, #0F1729, #131E35);
    border: 1px solid #1E3A5F; border-radius: 16px; padding: 22px 20px; margin-bottom: 4px;
}
.card-title { font-size:13px; font-weight:600; color:#64B5F6 !important;
    letter-spacing:0.8px; text-transform:uppercase; margin-bottom:10px; }
.file-ok { display:inline-flex; align-items:center; gap:8px;
    background:rgba(34,197,94,0.12); border:1px solid rgba(34,197,94,0.3);
    border-radius:8px; padding:6px 12px; font-size:12px;
    color:#22C55E !important; margin-top:8px; font-weight:500; }
div[data-testid="stButton"] > button {
    background:linear-gradient(135deg,#2563EB,#7C3AED) !important;
    color:white !important; border:none !important; border-radius:12px !important;
    padding:14px 40px !important; font-size:16px !important; font-weight:600 !important;
    width:100% !important; box-shadow:0 4px 20px rgba(37,99,235,0.35) !important; }
div[data-testid="stButton"] > button:disabled {
    background:linear-gradient(135deg,#1e3a5f,#2d1b5e) !important; color:#4B5563 !important; }
.step-container { display:flex; flex-direction:column; gap:10px; padding:20px;
    background:#0D1526; border-radius:14px; border:1px solid #1E3A5F; margin:16px 0; }
.step-row { display:flex; align-items:center; gap:14px; padding:10px 14px;
    border-radius:8px; font-size:14px; font-weight:500; }
.step-row.done    { background:rgba(34,197,94,0.08);  color:#4ADE80 !important; }
.step-row.running { background:rgba(59,130,246,0.12); color:#60A5FA !important; }
.step-row.pending { color:#4B5563 !important; }
.step-icon { font-size:16px; min-width:22px; }
.metric-grid { display:grid; grid-template-columns:repeat(4,1fr); gap:16px; margin:20px 0; }
.metric-tile { background:linear-gradient(145deg,#0F1729,#131E35);
    border:1px solid #1E3A5F; border-radius:14px; padding:20px; text-align:center; }
.metric-val { font-size:32px; font-weight:700; color:#60A5FA !important; }
.metric-lbl { font-size:12px; color:#64748B !important; margin-top:4px; font-weight:500; }
[data-testid="stDownloadButton"] > button {
    background:linear-gradient(135deg,#059669,#0D9488) !important;
    color:white !important; border:none !important; border-radius:12px !important;
    padding:14px 40px !important; font-size:16px !important; font-weight:600 !important;
    width:100% !important; box-shadow:0 4px 20px rgba(5,150,105,0.35) !important; margin-top:8px; }
.log-area { background:#060B18; border:1px solid #1E3A5F; border-radius:10px;
    padding:16px 20px; font-family:monospace !important; font-size:12px; color:#475569;
    max-height:340px; overflow-y:auto; line-height:1.8; }
.log-area .ok   { color:#22C55E; } .log-area .warn { color:#F59E0B; }
.log-area .err  { color:#EF4444; } .log-area .info { color:#60A5FA; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div style="background:linear-gradient(135deg,#0F1729 0%,#131E35 50%,#0F1729 100%);
     border:1px solid #1E3A5F; border-radius:20px; padding:36px 40px 28px; margin-bottom:28px;">
  <div style="display:flex;align-items:center;gap:18px;margin-bottom:10px;">
    <div style="background:linear-gradient(135deg,#2563EB,#7C3AED);border-radius:14px;
         padding:12px;font-size:26px;line-height:1;">📊</div>
    <div>
      <h1 style="margin:0;font-size:26px;font-weight:700;">KRA Auto-Updater</h1>
      <p style="margin:4px 0 0;color:#64748B;font-size:14px;">IFB Service · Cochin Cluster · Universal Executive KRA</p>
    </div>
  </div>
  <p style="color:#94A3B8;font-size:14px;margin:0;">
    Works for any executive — auto-detects franchises, maps correct rows per tab regardless of order.
  </p>
</div>
""", unsafe_allow_html=True)

# ── CONSTANTS ──────────────────────────────────────────────────────────────────
MONTHS = ["April","May","June","July","August","September",
          "October","November","December","January","February","March"]

def fy_idx(m):          return MONTHS.index(m)
def kra_col(m):         return 4 + fy_idx(m)
def sub_col_letter(m):  return chr(ord('D') + fy_idx(m))
def cluster_month_col(m): return 4 + fy_idx(m)
def mc_closed_col(m):   return 6 + fy_idx(m) * 2
def amc_ew_base_col(m): return 5 + fy_idx(m) * 4

# Spare sheet (AMC Per Call Cost) layout per franchise row-group:
# col1=FrCode, col2=Exec, col3=FrName,
# then per month: ZWR, AMC_Calls, Cost_Per_Call  → 3 cols per month
# base month col = 4 + fy_idx * 3
def spare_amc_col(m):
    # AMC spare: ZWR = base, Calls = base+1, CostPerCall = base+2
    return 4 + fy_idx(m) * 3

# SA Attendance sheet layout per franchise:
# col1=FrCode, col2=Exec, ...
# month col = total SA count column, and the % is a stored value (not formula)
# Structure confirmed: col = 4 + fy_idx (same as cluster_month_col)
# BUT SA Attendance has: col A = FrCode, col B = Executive, col C = Franchise name,
# then Apr=col4, May=col5 ... (stores % attendance as a decimal or integer)

def norm(s):
    return re.sub(r'[^a-z0-9]', '', str(s or "").lower())

# ── FILE UPLOADS ───────────────────────────────────────────────────────────────
st.markdown('<p style="font-size:13px;color:#64748B;font-weight:600;letter-spacing:0.8px;'
            'text-transform:uppercase;margin-bottom:12px;">① Upload Input Files</p>',
            unsafe_allow_html=True)
c1, c2, c3 = st.columns(3)
with c1:
    st.markdown('<div class="upload-card"><p class="card-title">📁 KL Cluster Report</p>', unsafe_allow_html=True)
    kl_file = st.file_uploader("kl", type=["xlsx"], key="kl", label_visibility="collapsed")
    if kl_file: st.markdown(f'<div class="file-ok">✓ &nbsp;{kl_file.name}</div>', unsafe_allow_html=True)
    else: st.markdown('<p style="color:#374151;font-size:12px;margin-top:4px;">MC Reg · AMC+EW · SA Attendance · Spare · Social · ESS · ACC</p>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
with c2:
    st.markdown('<div class="upload-card"><p class="card-title">📁 Parameter Dashboard</p>', unsafe_allow_html=True)
    param_file = st.file_uploader("param", type=["xlsx"], key="param", label_visibility="collapsed")
    if param_file: st.markdown(f'<div class="file-ok">✓ &nbsp;{param_file.name}</div>', unsafe_allow_html=True)
    else: st.markdown('<p style="color:#374151;font-size:12px;margin-top:4px;">INS · SER · CSS · NR · MC Hit · Rep Calls · SA Prod</p>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
with c3:
    st.markdown('<div class="upload-card"><p class="card-title">📁 KRA Template</p>', unsafe_allow_html=True)
    kra_file = st.file_uploader("kra", type=["xlsx"], key="kra", label_visibility="collapsed")
    if kra_file: st.markdown(f'<div class="file-ok">✓ &nbsp;{kra_file.name}</div>', unsafe_allow_html=True)
    else: st.markdown('<p style="color:#374151;font-size:12px;margin-top:4px;">Any executive\'s KRA template (14+ tabs)</p>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

# ── MONTH SELECTOR ─────────────────────────────────────────────────────────────
st.markdown('<p style="font-size:13px;color:#64748B;font-weight:600;letter-spacing:0.8px;'
            'text-transform:uppercase;margin:20px 0 10px;">② Select Month</p>',
            unsafe_allow_html=True)
mc1, mc2 = st.columns([1, 3])
with mc1:
    selected_month = st.selectbox("Month", MONTHS, index=9, label_visibility="collapsed")
with mc2:
    fi   = fy_idx(selected_month); kc = kra_col(selected_month)
    scl  = sub_col_letter(selected_month)
    mcc  = mc_closed_col(selected_month); cmc = cluster_month_col(selected_month)
    amc_base = amc_ew_base_col(selected_month)
    spare_col = spare_amc_col(selected_month)
    st.markdown(f"""
    <div style="background:#0F1729;border:1px solid #1E3A5F;border-radius:12px;
         padding:14px 20px;display:flex;gap:24px;align-items:center;flex-wrap:wrap;">
      <div><div style="font-size:11px;color:#475569;text-transform:uppercase;">Month</div>
           <div style="font-size:18px;font-weight:700;color:#60A5FA;">{selected_month}</div></div>
      <div><div style="font-size:11px;color:#475569;text-transform:uppercase;">FY Index</div>
           <div style="font-size:18px;font-weight:700;color:#A78BFA;">{fi+1}/12</div></div>
      <div><div style="font-size:11px;color:#475569;text-transform:uppercase;">Cluster Col</div>
           <div style="font-size:18px;font-weight:700;color:#F472B6;">col {cmc}</div></div>
      <div><div style="font-size:11px;color:#475569;text-transform:uppercase;">MC Closed Col</div>
           <div style="font-size:18px;font-weight:700;color:#FB923C;">col {mcc}</div></div>
      <div><div style="font-size:11px;color:#475569;text-transform:uppercase;">Spare Col</div>
           <div style="font-size:18px;font-weight:700;color:#34D399;">col {spare_col}</div></div>
      <div><div style="font-size:11px;color:#475569;text-transform:uppercase;">KRA Col</div>
           <div style="font-size:18px;font-weight:700;color:#60A5FA;">{kc} → {scl}</div></div>
    </div>""", unsafe_allow_html=True)

# ── PROCESS BUTTON ─────────────────────────────────────────────────────────────
st.markdown('<p style="font-size:13px;color:#64748B;font-weight:600;letter-spacing:0.8px;'
            'text-transform:uppercase;margin:20px 0 10px;">③ Process</p>',
            unsafe_allow_html=True)
all_ready = kl_file and param_file and kra_file
if not all_ready:
    missing = [x for x,f in [("KL Cluster",kl_file),("Parameter",param_file),("KRA Template",kra_file)] if not f]
    st.markdown(f'<div style="background:rgba(251,146,60,0.08);border:1px solid rgba(251,146,60,0.3);'
                f'border-radius:10px;padding:12px 18px;color:#FB923C;font-size:13px;margin-bottom:12px;">'
                f'⚠️ Still needed: {" · ".join(missing)}</div>', unsafe_allow_html=True)
btn = st.button("⚡  Generate Updated KRA", disabled=not all_ready)

# ── HELPERS ────────────────────────────────────────────────────────────────────
def safe_int(v):
    try:
        if v in (None, "") or str(v).strip() in ('nan','None','—','#N/A','#REF!'): return 0
        return int(float(v))
    except: return 0

def safe_float(v):
    try:
        if v in (None, "") or str(v).strip() in ('nan','None','—','#N/A','#REF!'): return 0.0
        return float(v)
    except: return 0.0

def find_sheet(wb, keywords):
    kws = [k.lower() for k in keywords]
    for name in wb.sheetnames:
        if all(k in name.lower() for k in kws): return wb[name]
    for name in wb.sheetnames:
        if any(k in name.lower() for k in kws): return wb[name]
    return None

def fuzzy_score(a, b):
    if not a or not b: return 0
    if a == b: return 1.0
    if a in b or b in a: return 0.9
    plen = 0
    for x, y in zip(a, b):
        if x == y: plen += 1
        else: break
    prefix_score = plen / max(len(a), len(b)) if plen >= 3 else 0
    at = set(re.findall(r'[a-z0-9]{3,}', a))
    bt = set(re.findall(r'[a-z0-9]{3,}', b))
    token_score = len(at & bt) / max(len(at), len(bt)) * 0.85 if at and bt else 0
    return max(prefix_score, token_score)

def build_code_lookup(ws, code_col, val_col, skip_rows=1):
    """Build {7-digit-code: value} from a sheet."""
    lkp = {}
    if not ws: return lkp
    for r in range(1 + skip_rows, ws.max_row + 1):
        code = str(ws.cell(r, code_col).value or "").strip()
        if re.match(r'^\d{7}$', code):
            lkp[code] = ws.cell(r, val_col).value
    return lkp

def scan_all_sheets(wb):
    """Print all sheet names for debugging."""
    return [s for s in wb.sheetnames]

def auto_detect_codes(kra_franchises, ws_ins):
    param_entries = []
    for r in range(3, ws_ins.max_row + 1):
        code = str(ws_ins.cell(r, 1).value or "").strip()
        name = str(ws_ins.cell(r, 2).value or "").strip()
        if re.match(r'^\d{7}$', code) and name:
            param_entries.append((norm(name), code, name))
    result = {}
    for kname in kra_franchises:
        kn = norm(kname)
        best = max(param_entries, key=lambda x: fuzzy_score(kn, x[0]), default=None)
        if best and fuzzy_score(kn, best[0]) >= 0.4:
            result[kname] = (best[1], best[2], round(fuzzy_score(kn, best[0]), 2))
        else:
            result[kname] = (None, None, 0)
    return result

def auto_build_tab_row_maps(ws_map, kra_franchises):
    """
    Per-tab independent row detection via col B scan.
    Handles any franchise ordering per sheet.
    """
    result = {}
    for tab, ws in ws_map.items():
        if not ws: result[tab] = {}; continue
        col_b_rows = []
        for r in range(1, ws.max_row + 1):
            b = str(ws.cell(r, 2).value or "").strip()
            if b and norm(b) not in ('franchisee','slno','cluster','executive',''):
                col_b_rows.append((r, b, norm(b)))

        tab_map = {}
        used_rows = set()
        for kname in kra_franchises:
            kn = norm(kname)
            best_row, best_score = None, 0
            for r, b, bn in col_b_rows:
                if r in used_rows: continue
                sc = fuzzy_score(kn, bn)
                if sc > best_score:
                    best_score = sc
                    best_row   = r
            if best_row and best_score >= 0.35:
                tab_map[kname] = best_row
                used_rows.add(best_row)

        # Detect overall / total row
        for r, b, bn in col_b_rows:
            if any(x in b.lower() for x in ["overall","total","grand"]):
                tab_map["OVERALL"] = r; break

        result[tab] = tab_map
    return result

# ── SA ATTENDANCE lookup ──────────────────────────────────────────────────────
def build_sa_attendance_lookup(ws, month_col):
    """
    SA Attendance sheet:
      Col A = Franchise Code (7-digit)
      Col B = Executive Name
      Col C = Franchise Name  (some sheets skip col C and put data from col B)
      Month cols = 4 + fy_idx (April=4, May=5 ... March=15)
    
    Stores the ATTENDANCE % directly (e.g. 94.8 meaning 94.8%)
    We also need total No. of SAs. SA Attendance sheet likely has:
      Row per franchise with attendance %.
    We return {code: attendance_pct}
    """
    lkp = {}
    if not ws: return lkp
    for r in range(1, ws.max_row + 1):
        code = str(ws.cell(r, 1).value or "").strip()
        if re.match(r'^\d{7}$', code):
            val = ws.cell(r, month_col).value
            lkp[code] = safe_float(val)
    return lkp

# ── SPARE (AMC Per Call Cost) lookup ─────────────────────────────────────────
def build_spare_lookups(ws, month_col):
    """
    AMC Per Call Cost sheet structure per row:
      Col 1 = Franchise Code (7-digit) — BUT many rows share same code (AMC row, WTY row)
      Col 2 = Executive
      Col 3 = Franchise Name
      Then per month: ZWR_value, Calls, Cost_Per_Call  (3 cols per month)
      
    BUT looking at actual data, the sheet has TWO separate sections:
      1) AMC section: rows with FrCode, then ZWR(spare spend), AMC calls, cost/call
      2) WTY section: same codes, WTY ZWR, WTY calls, WTY cost/call
    
    We distinguish AMC vs WTY rows by scanning adjacent text or row position.
    Simpler approach: scan for duplicate codes and assign first = AMC, second = WTY.
    
    Returns {code: {amc_zwr, amc_calls, wty_zwr, wty_calls}}
    """
    lkp = {}
    if not ws: return lkp
    
    code_seen = {}  # code -> count of times seen
    for r in range(1, ws.max_row + 1):
        code = str(ws.cell(r, 1).value or "").strip()
        if not re.match(r'^\d{7}$', code): continue
        
        zwr_val   = safe_float(ws.cell(r, month_col).value)
        calls_val = safe_int(ws.cell(r, month_col + 1).value)
        
        if code not in lkp:
            lkp[code] = {"amc_zwr": 0.0, "amc_calls": 0, "wty_zwr": 0.0, "wty_calls": 0}
            code_seen[code] = 0
        
        code_seen[code] += 1
        if code_seen[code] == 1:
            # First occurrence = AMC row
            lkp[code]["amc_zwr"]   = zwr_val
            lkp[code]["amc_calls"] = calls_val
        elif code_seen[code] == 2:
            # Second occurrence = WTY row
            lkp[code]["wty_zwr"]   = zwr_val
            lkp[code]["wty_calls"] = calls_val
    
    return lkp

def wr(ws, r, c, v):
    if ws and r and c:
        try: ws.cell(row=r, column=c, value=v)
        except: pass

# ── MAIN PROCESSING ────────────────────────────────────────────────────────────
if btn and all_ready:
    logs = []
    def log(msg, t="info"):
        icon = {"ok":"✅","warn":"⚠️","err":"❌","info":"→"}.get(t,"→")
        logs.append(f'<span class="{t}">{icon} {msg}</span>')

    steps_ph = st.empty()
    STEPS = [
        ("📂","Reading source files"),
        ("🔍","Auto-detecting franchise codes"),
        ("📊","Building all lookups"),
        ("🗺️","Building per-tab row maps"),
        ("✍️","Writing data to all KRA tabs"),
        ("🔗","Wiring KRA Sheet dashboard"),
        ("💾","Saving & finalizing"),
    ]
    def render_steps(cur):
        html = '<div class="step-container">'
        for i,(icon,label) in enumerate(STEPS):
            if i < cur:  cls,tick = "done","✓"
            elif i==cur: cls,tick = "running","◉"
            else:        cls,tick = "pending","○"
            html += f'<div class="step-row {cls}"><span class="step-icon">{tick}</span><span>{icon} {label}</span></div>'
        return html + '</div>'

    try:
        # ── STEP 0: Read files ──────────────────────────────────────────────
        steps_ph.markdown(render_steps(0), unsafe_allow_html=True)
        fi       = fy_idx(selected_month)
        kc       = kra_col(selected_month)
        scl      = sub_col_letter(selected_month)
        cmc      = cluster_month_col(selected_month)
        mcc      = mc_closed_col(selected_month)
        amc_base = amc_ew_base_col(selected_month)
        sp_col   = spare_amc_col(selected_month)

        kl_wb    = load_workbook(BytesIO(kl_file.read()),    data_only=True)
        param_wb = load_workbook(BytesIO(param_file.read()), data_only=True)
        kra_wb   = load_workbook(BytesIO(kra_file.read()))

        kl_sheets    = scan_all_sheets(kl_wb)
        param_sheets = scan_all_sheets(param_wb)
        kra_sheets   = scan_all_sheets(kra_wb)
        log(f"KL sheets: {kl_sheets}", "info")
        log(f"Param sheets: {param_sheets}", "info")
        log(f"KRA sheets: {kra_sheets}", "info")
        log(f"Month={selected_month} | cluster_col={cmc} | mc_col={mcc} | spare_col={sp_col} | kra_col={kc}({scl})", "ok")

        # ── STEP 1: Read KRA franchises + auto-detect codes ─────────────────
        steps_ph.markdown(render_steps(1), unsafe_allow_html=True)
        ws_cl_kra = find_sheet(kra_wb, ["call load"])
        kra_franchises = []
        if ws_cl_kra:
            for r in range(1, ws_cl_kra.max_row + 1):
                b = str(ws_cl_kra.cell(r,2).value or "").strip()
                c_val = str(ws_cl_kra.cell(r,3).value or "").strip()
                if b and c_val.lower() == "installation" and b.lower() not in ("franchisee",""):
                    kra_franchises.append(b)

        ws_ins   = find_sheet(param_wb, ["ins"])
        code_map = auto_detect_codes(kra_franchises, ws_ins)
        KRA_CODES = {k: v[0] for k, v in code_map.items()}
        all_names = kra_franchises + ["OVERALL"]

        for kname, (code, matched, score) in code_map.items():
            if code: log(f"'{kname}' → {code} '{matched}' (score:{score})", "ok")
            else:    log(f"'{kname}' → NO CODE", "warn")
        log(f"{len(kra_franchises)} franchises | {sum(1 for v in KRA_CODES.values() if v)} codes resolved", "ok")

        # ── STEP 2: Build lookups ───────────────────────────────────────────
        steps_ph.markdown(render_steps(2), unsafe_allow_html=True)

        # --- Parameter Dashboard sheets ---
        ws_nr     = find_sheet(param_wb, ["nr"])
        ws_css    = find_sheet(param_wb, ["css"])
        ws_mc_hit = find_sheet(param_wb, ["mc hit"])
        ws_rep    = find_sheet(param_wb, ["rep call","rep calls"])
        ws_sa_p   = find_sheet(param_wb, ["sa prod"])

        # --- KL Cluster Report sheets ---
        ws_mc_reg  = find_sheet(kl_wb, ["mc reg"])
        ws_abv2    = find_sheet(kl_wb, ["abv 2"])
        ws_sa_att  = find_sheet(kl_wb, ["sa attend"])
        ws_social  = find_sheet(kl_wb, ["social"])
        ws_ess_cl  = find_sheet(kl_wb, ["ess bdg"])
        ws_acc_cl  = find_sheet(kl_wb, ["acc bdg"])
        ws_dhukhan = find_sheet(kl_wb, ["apni"])
        ws_amc_ew  = find_sheet(kl_wb, ["amc+ew","amc ew","amc+"])
        ws_spare   = find_sheet(kl_wb, ["per call cost","amc per call","spare"])

        log(f"SA Attend sheet: '{ws_sa_att.title if ws_sa_att else 'NOT FOUND'}'", "ok" if ws_sa_att else "warn")
        log(f"Spare sheet: '{ws_spare.title if ws_spare else 'NOT FOUND'}'", "ok" if ws_spare else "warn")

        # Inspect SA Attendance sheet header to confirm column layout
        if ws_sa_att:
            header_row = [str(ws_sa_att.cell(1, c).value or "") for c in range(1, 20)]
            log(f"SA Attend row1: {header_row}", "info")
            # Check first few data rows
            for r in range(2, 6):
                row_data = [str(ws_sa_att.cell(r, c).value or "") for c in range(1, 8)]
                log(f"SA Attend row{r}: {row_data}", "info")

        # Inspect Spare sheet header
        if ws_spare:
            header_row = [str(ws_spare.cell(1, c).value or "") for c in range(1, 16)]
            log(f"Spare row1: {header_row}", "info")
            for r in range(2, 6):
                row_data = [str(ws_spare.cell(r, c).value or "") for c in range(1, 10)]
                log(f"Spare row{r}: {row_data}", "info")

        # ── Standard code-keyed lookups ────────────────────────────────────
        lk_ins_closed = build_code_lookup(ws_ins,    1, 3, skip_rows=2)
        lk_ins_6hrs   = build_code_lookup(ws_ins,    1, 4, skip_rows=2)
        lk_ser_closed = build_code_lookup(ws_ins,    1, 5, skip_rows=2)
        lk_ser_24hrs  = build_code_lookup(ws_ins,    1, 6, skip_rows=2)
        lk_nr_total   = build_code_lookup(ws_nr,     1, 9,  skip_rows=1)
        lk_nr_closed  = build_code_lookup(ws_nr,     1, 10, skip_rows=1)
        lk_css_ok     = build_code_lookup(ws_css,    1, 3, skip_rows=1)
        lk_css_not_ok = build_code_lookup(ws_css,    1, 4, skip_rows=1)
        lk_css_happy  = build_code_lookup(ws_css,    1, 5, skip_rows=1)
        lk_mc_hit_reg = build_code_lookup(ws_mc_hit, 1, 3, skip_rows=1)
        lk_mc_hit_cl  = build_code_lookup(ws_mc_hit, 1, 4, skip_rows=1)
        lk_rep_total  = build_code_lookup(ws_rep,    2, 6, skip_rows=1)
        lk_rep_ticket = build_code_lookup(ws_rep,    2, 7, skip_rows=1)

        # SA Prod: No. of SAs = col 10
        lk_sa_count = build_code_lookup(ws_sa_p, 2, 10, skip_rows=1)

        # SA Attendance: attendance % stored directly per month column
        # col1=code, month_col=cmc (same layout as other KL sheets)
        lk_sa_att_pct = build_sa_attendance_lookup(ws_sa_att, cmc)

        # KL Cluster lookups
        lk_mc_closed   = build_code_lookup(ws_mc_reg,  1, mcc, skip_rows=2)
        lk_abv2        = build_code_lookup(ws_abv2,    1, cmc, skip_rows=1)
        lk_social      = build_code_lookup(ws_social,  1, cmc, skip_rows=1)
        lk_ess_tgt     = build_code_lookup(ws_ess_cl,  1, 22, skip_rows=1)
        lk_ess_ach     = build_code_lookup(ws_ess_cl,  1, 23, skip_rows=1)
        lk_acc_tgt     = build_code_lookup(ws_acc_cl,  1, 22, skip_rows=1)
        lk_acc_ach     = build_code_lookup(ws_acc_cl,  1, 23, skip_rows=1)
        lk_dhukhan     = build_code_lookup(ws_dhukhan, 1, cmc, skip_rows=1)

        # AMC+EW per month block
        lk_amc_bdg_nos = build_code_lookup(ws_amc_ew, 1, amc_base,   skip_rows=2)
        lk_amc_bdg_val = build_code_lookup(ws_amc_ew, 1, amc_base+1, skip_rows=2)
        lk_amc_nos_ach = build_code_lookup(ws_amc_ew, 1, amc_base+2, skip_rows=2)
        lk_amc_val_ach = build_code_lookup(ws_amc_ew, 1, amc_base+3, skip_rows=2)

        # Spare (AMC Per Call Cost) lookups
        lk_spare = build_spare_lookups(ws_spare, sp_col)

        # Log sample spare data
        sample_codes = list(lk_spare.keys())[:3]
        for sc in sample_codes:
            log(f"Spare sample {sc}: {lk_spare[sc]}", "info")

        log(f"All lookups built | SA Att entries:{len(lk_sa_att_pct)} | Spare entries:{len(lk_spare)}", "ok")

        # ── Build data dict ──────────────────────────────────────────────────
        def g(lk, kname, fn=safe_int):
            code = KRA_CODES.get(kname)
            return fn(lk.get(code, 0)) if code else fn(0)

        def g_spare(kname, field):
            code = KRA_CODES.get(kname)
            if not code: return 0
            return lk_spare.get(code, {}).get(field, 0)

        data = {}
        for kname in all_names:
            ins_c  = g(lk_ins_closed, kname)
            ser_c  = g(lk_ser_closed, kname)
            mc_c   = g(lk_mc_hit_cl,  kname)

            # SA Attendance:
            #   sa_total    = No. of SAs from SA Prod sheet
            #   sa_att_pct  = Attendance % from SA Attendance sheet (stored value)
            #   sa_25days   = round(sa_total × sa_att_pct / 100)  ← SAs who attended 25 days
            sa_total   = g(lk_sa_count, kname)
            sa_att_pct = g(lk_sa_att_pct, kname, safe_float)
            sa_25days  = round(sa_total * sa_att_pct / 100) if sa_att_pct > 1 else round(sa_total * sa_att_pct)

            data[kname] = {
                # Call Load
                "ins_closed":    ins_c,
                "ins_6hrs":      g(lk_ins_6hrs,   kname),
                "ser_closed":    ser_c,
                "ser_24hrs":     g(lk_ser_24hrs,   kname),
                "mc_hit_closed": mc_c,
                "mc_hit_reg":    g(lk_mc_hit_reg,  kname),
                # >2 days
                "ser_closed_2d": ser_c,
                "avg_pend":      round(g(lk_abv2, kname, safe_float), 2),
                # Repeat
                "rep_ticket":    g(lk_rep_ticket,  kname),
                "rep_total":     g(lk_rep_total,   kname),
                # CSS
                "css_ok":        g(lk_css_ok,      kname),
                "css_not_ok":    g(lk_css_not_ok,  kname),
                "css_happy":     g(lk_css_happy,   kname),
                # NR
                "nr_closed":     g(lk_nr_closed,   kname),
                "nr_total":      g(lk_nr_total,    kname),
                # Social
                "total_calls":   ins_c + ser_c + mc_c,
                "social":        g(lk_social,      kname),
                # SA Attendance
                "sa_total":      sa_total,
                "sa_att_pct":    sa_att_pct,
                "sa_25days":     sa_25days,
                # AMC Achievement
                "amc_bdg_nos":   g(lk_amc_bdg_nos, kname, safe_float),
                "amc_nos_ach":   g(lk_amc_nos_ach, kname, safe_float),
                "amc_bdg_val":   g(lk_amc_bdg_val, kname, safe_float),
                "amc_val_ach":   g(lk_amc_val_ach, kname, safe_float),
                # Essential / Accessories
                "ess_tgt":       g(lk_ess_tgt,  kname, safe_float),
                "ess_ach":       g(lk_ess_ach,  kname, safe_float),
                "acc_tgt":       g(lk_acc_tgt,  kname, safe_float),
                "acc_ach":       g(lk_acc_ach,  kname, safe_float),
                # Exchange
                "exchange":      g(lk_dhukhan,  kname),
                # Spare Consumption (AMC Per Call Cost)
                "spare_amc_zwr":   safe_float(g_spare(kname, "amc_zwr")),
                "spare_amc_calls": safe_int(g_spare(kname, "amc_calls")),
                "spare_wty_zwr":   safe_float(g_spare(kname, "wty_zwr")),
                "spare_wty_calls": safe_int(g_spare(kname, "wty_calls")),
            }

        log(f"Data built for {len(data)} entries", "ok")

        # ── STEP 3: Build per-tab row maps ───────────────────────────────────
        steps_ph.markdown(render_steps(3), unsafe_allow_html=True)
        ws_map = {
            "Call Load":         find_sheet(kra_wb, ["call load"]),
            "Installation":      find_sheet(kra_wb, ["installation"]),
            "Service":           find_sheet(kra_wb, ["service"]),
            ">2 days Pending":   find_sheet(kra_wb, [">2"]),
            "CSS":               find_sheet(kra_wb, ["css"]),
            "Negative Response": find_sheet(kra_wb, ["negative"]),
            "Social M Calls":    find_sheet(kra_wb, ["social"]),
            "Repeat Calls":      find_sheet(kra_wb, ["repeat"]),
            "MC Calls":          find_sheet(kra_wb, ["mc call"]),
            "SA Attendance":     find_sheet(kra_wb, ["sa attend"]),
            "AMC Achievement":   find_sheet(kra_wb, ["amc achiev"]),
            "Essential Budget":  find_sheet(kra_wb, ["essential"]),
            "Accesories Budget": find_sheet(kra_wb, ["accesories","accessories"]),
            "Exchange":          find_sheet(kra_wb, ["exchange"]),
            "Spare Consumption": find_sheet(kra_wb, ["spare","cosnumption","consumption"]),
        }
        missing_ws = [k for k,v in ws_map.items() if not v]
        if missing_ws: log(f"Missing KRA sheets: {missing_ws}", "warn")

        kra_rmaps = auto_build_tab_row_maps(ws_map, all_names)

        for tab, rmap in kra_rmaps.items():
            matched   = [k for k in kra_franchises if k in rmap]
            unmatched = [k for k in kra_franchises if k not in rmap]
            if unmatched: log(f"'{tab}': unmatched → {unmatched}", "warn")
            else:         log(f"'{tab}': all {len(matched)} matched ✓", "ok")

        # Debug: show detected rows for SA Attendance and Spare
        for tab in ["SA Attendance", "Spare Consumption"]:
            rmap = kra_rmaps.get(tab, {})
            log(f"Row map '{tab}': {rmap}", "info")

        # ── STEP 4: Write data ───────────────────────────────────────────────
        steps_ph.markdown(render_steps(4), unsafe_allow_html=True)
        updated = []
        for kname in all_names:
            d = data[kname]
            for tab, ws in ws_map.items():
                if not ws: continue
                sr = kra_rmaps.get(tab, {}).get(kname)
                if not sr: continue
                col = kc

                if tab == "Call Load":
                    wr(ws, sr,   col, d["ins_closed"])
                    wr(ws, sr+1, col, d["ser_closed"])
                    wr(ws, sr+2, col, d["mc_hit_closed"])

                elif tab == "Installation":
                    wr(ws, sr,   col, d["ins_closed"])
                    wr(ws, sr+1, col, d["ins_6hrs"])

                elif tab == "Service":
                    wr(ws, sr,   col, d["ser_closed"])
                    wr(ws, sr+1, col, d["ser_24hrs"])

                elif tab == ">2 days Pending":
                    wr(ws, sr,   col, d["ser_closed_2d"])
                    wr(ws, sr+1, col, d["avg_pend"])

                elif tab == "Repeat Calls":
                    wr(ws, sr,   col, d["rep_ticket"])
                    wr(ws, sr+1, col, d["rep_total"])

                elif tab == "SA Attendance":
                    # KRA SA Attendance tab block per franchise:
                    # sr+0: Total No of SA      ← sa_total
                    # sr+1: SA 25 Days count    ← sa_25days (= sa_total × att% / 100)
                    # sr+2: SA Attendance %     ← FORMULA (=sr+1/sr+0*100) → DO NOT TOUCH
                    wr(ws, sr,   col, d["sa_total"])
                    wr(ws, sr+1, col, d["sa_25days"])

                elif tab == "CSS":
                    wr(ws, sr,   col, d["css_happy"])
                    wr(ws, sr+1, col, d["css_ok"])
                    wr(ws, sr+2, col, d["css_not_ok"])

                elif tab == "Negative Response":
                    wr(ws, sr,   col, d["nr_closed"])
                    wr(ws, sr+1, col, d["nr_total"])

                elif tab == "Social M Calls":
                    wr(ws, sr,   col, d["total_calls"])
                    wr(ws, sr+1, col, d["social"])

                elif tab == "MC Calls":
                    wr(ws, sr,   col, d["mc_hit_closed"])
                    wr(ws, sr+1, col, d["mc_hit_reg"])

                elif tab == "AMC Achievement":
                    # sr+0: AMC Target Nos  → BDG Nos
                    # sr+1: AMC Nos Ach     → AMC Nos achieved
                    # sr+2: AMC Nos %       → FORMULA
                    # sr+3: AMC Value Tgt   → BDG Value
                    # sr+4: AMC Value Ach   → AMC Value achieved
                    # sr+5: AMC Value %     → FORMULA
                    wr(ws, sr,   col, d["amc_bdg_nos"])
                    wr(ws, sr+1, col, d["amc_nos_ach"])
                    wr(ws, sr+3, col, d["amc_bdg_val"])
                    wr(ws, sr+4, col, d["amc_val_ach"])

                elif tab == "Essential Budget":
                    wr(ws, sr,   col, d["ess_tgt"])
                    wr(ws, sr+1, col, d["ess_ach"])

                elif tab == "Accesories Budget":
                    wr(ws, sr,   col, d["acc_tgt"])
                    wr(ws, sr+1, col, d["acc_ach"])

                elif tab == "Exchange":
                    wr(ws, sr, col, d["exchange"])

                elif tab == "Spare Consumption":
                    # KRA Spare Consumption tab block per franchise:
                    # sr+0: AMC ZWR (Spare spend on AMC calls)  ← spare_amc_zwr
                    # sr+1: AMC Calls Closed                    ← spare_amc_calls
                    # sr+2: AMC Cost Per Call                   ← FORMULA (=sr+0/sr+1) → DO NOT TOUCH
                    # sr+3: WTY ZWR (Spare spend on Wty calls)  ← spare_wty_zwr
                    # sr+4: WTY Calls Closed                    ← spare_wty_calls
                    # sr+5: WTY Cost Per Call                   ← FORMULA → DO NOT TOUCH
                    wr(ws, sr,   col, d["spare_amc_zwr"])
                    wr(ws, sr+1, col, d["spare_amc_calls"])
                    wr(ws, sr+3, col, d["spare_wty_zwr"])
                    wr(ws, sr+4, col, d["spare_wty_calls"])

                if tab not in updated: updated.append(tab)

        log(f"Written to {len(updated)} tabs", "ok")

        # ── STEP 5: Wire dashboard ───────────────────────────────────────────
        steps_ph.markdown(render_steps(5), unsafe_allow_html=True)
        ws_dash = find_sheet(kra_wb, ["kra sheet"])
        if ws_dash:
            c = scl
            formulas = {
                9:  f"='Call Load'!{c}59",
                10: f"='Call Load'!{c}60",
                11: f"='Call Load'!{c}61",
                12: f"='Call Load'!{c}62",
                13: f"=Installation!{c}39",
                14: f"=Service!{c}39",
                15: f"='>2 days Pending'!{c}39",
                16: f"=CSS!{c}50",
                17: f"='Negative Response'!{c}39",
                18: f"='Social M Calls'!{c}38",
                20: f"='Repeat Calls'!{c}39",
                21: f"='MC Calls'!{c}39",
                24: f"='AMC Achievement'!{c}72",
                25: f"='AMC Achievement'!{c}75",
                26: f"='AMC HIT Rate'!{c}78",
                27: f"='AMC HIT Rate'!{c}81",
                28: f"='Essential Budget'!{c}45",
                29: f"='Accesories Budget'!{c}48",
                30: f"='Spare Cosnumption'!{c}104",
            }
            for row_num, formula in formulas.items():
                ws_dash.cell(row=row_num, column=kra_col(selected_month)).value = formula
            log(f"Dashboard wired: {len(formulas)} rows → col {c}", "ok")
        else:
            log("KRA Sheet tab not found", "warn")

        # ── STEP 6: Save ─────────────────────────────────────────────────────
        steps_ph.markdown(render_steps(6), unsafe_allow_html=True)
        out = BytesIO(); kra_wb.save(out); out.seek(0)
        log("Saved ✓", "ok")
        steps_ph.markdown(render_steps(7), unsafe_allow_html=True)

        # ── RESULTS ──────────────────────────────────────────────────────────
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,rgba(34,197,94,0.1),rgba(5,150,105,0.05));
             border:1px solid rgba(34,197,94,0.3);border-radius:16px;padding:20px 24px;
             margin:20px 0;display:flex;align-items:center;gap:14px;">
          <div style="font-size:28px;">🎉</div>
          <div>
            <div style="font-size:17px;font-weight:700;color:#4ADE80;">KRA Successfully Updated!</div>
            <div style="font-size:13px;color:#6EE7B7;margin-top:3px;">
              {selected_month} 2026 &nbsp;·&nbsp; {len(kra_franchises)} franchises &nbsp;·&nbsp;
              {len(updated)} tabs written
            </div>
          </div>
        </div>""", unsafe_allow_html=True)

        ov = data.get("OVERALL", {})
        st.markdown(f"""
        <div class="metric-grid">
          <div class="metric-tile"><div class="metric-val">{len(kra_franchises)}</div><div class="metric-lbl">Franchises</div></div>
          <div class="metric-tile"><div class="metric-val">{len(updated)}</div><div class="metric-lbl">Tabs Updated</div></div>
          <div class="metric-tile"><div class="metric-val">{ov.get('ins_closed',0)}</div><div class="metric-lbl">INS Closed</div></div>
          <div class="metric-tile"><div class="metric-val">{ov.get('ser_closed',0)}</div><div class="metric-lbl">SER Closed</div></div>
        </div>""", unsafe_allow_html=True)

        # Code detection table
        st.markdown('<p style="font-size:13px;color:#64748B;font-weight:600;letter-spacing:0.8px;'
                    'text-transform:uppercase;margin:20px 0 8px;">Franchise Code Detection</p>',
                    unsafe_allow_html=True)
        st.dataframe(pd.DataFrame([{
            "KRA Franchise":      kname,
            "Detected Code":      code_map[kname][0] or "—",
            "Matched Param Name": code_map[kname][1] or "NOT FOUND",
            "Score":              f"{code_map[kname][2]:.2f}",
            "Status":             "✅" if code_map[kname][0] else "⚠️"
        } for kname in kra_franchises]), use_container_width=True, height=320)

        # Data preview
        st.markdown('<p style="font-size:13px;color:#64748B;font-weight:600;letter-spacing:0.8px;'
                    'text-transform:uppercase;margin:20px 0 8px;">SA + Spare Preview</p>',
                    unsafe_allow_html=True)
        st.dataframe(pd.DataFrame([{
            "Franchise":        kname,
            "SA Total":         data[kname]["sa_total"],
            "SA Att %":         f'{data[kname]["sa_att_pct"]:.1f}',
            "SA 25d Count":     data[kname]["sa_25days"],
            "AMC ZWR":          f'{data[kname]["spare_amc_zwr"]:,.0f}',
            "AMC Calls":        data[kname]["spare_amc_calls"],
            "WTY ZWR":          f'{data[kname]["spare_wty_zwr"]:,.0f}',
            "WTY Calls":        data[kname]["spare_wty_calls"],
        } for kname in kra_franchises]), use_container_width=True, height=360)

        # Full data preview
        st.markdown('<p style="font-size:13px;color:#64748B;font-weight:600;letter-spacing:0.8px;'
                    'text-transform:uppercase;margin:20px 0 8px;">Full Data Preview</p>',
                    unsafe_allow_html=True)
        st.dataframe(pd.DataFrame([{
            "Franchise":      kname,
            "INS":            data[kname]["ins_closed"],
            "SER":            data[kname]["ser_closed"],
            "MC":             data[kname]["mc_hit_closed"],
            "AMC Tgt Nos":    data[kname]["amc_bdg_nos"],
            "AMC Nos Ach":    data[kname]["amc_nos_ach"],
            "ESS Tgt":        f'{data[kname]["ess_tgt"]:,.0f}',
            "ESS Ach":        f'{data[kname]["ess_ach"]:,.0f}',
            "Exchange":       data[kname]["exchange"],
        } for kname in kra_franchises]), use_container_width=True, height=360)

        # Log
        st.markdown('<p style="font-size:13px;color:#64748B;font-weight:600;letter-spacing:0.8px;'
                    'text-transform:uppercase;margin:20px 0 8px;">Processing Log</p>',
                    unsafe_allow_html=True)
        st.markdown(f'<div class="log-area">{"<br>".join(logs)}</div>', unsafe_allow_html=True)

        # Download
        st.markdown('<p style="font-size:13px;color:#64748B;font-weight:600;letter-spacing:0.8px;'
                    'text-transform:uppercase;margin:20px 0 8px;">④ Download</p>',
                    unsafe_allow_html=True)
        st.download_button(
            label=f"⬇️  Download KRA — {selected_month} 2026",
            data=out,
            file_name=f"KRA-{selected_month}-2026.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        import traceback
        steps_ph.empty()
        st.error(f"❌ {e}")
        st.code(traceback.format_exc())

if not btn and not all_ready:
    st.markdown("""
    <div style="background:linear-gradient(145deg,#0F1729,#131E35);
         border:1px dashed #1E3A5F;border-radius:20px;padding:60px 40px;text-align:center;margin-top:8px;">
      <div style="font-size:48px;margin-bottom:16px;">📋</div>
      <h3 style="color:#4B5563;font-weight:600;margin:0 0 10px;">Ready to Begin</h3>
      <p style="color:#374151;font-size:14px;max-width:480px;margin:0 auto;line-height:1.7;">
        Upload all 3 files, select the month, and click Generate.
      </p>
    </div>""", unsafe_allow_html=True)
